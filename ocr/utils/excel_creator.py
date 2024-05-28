from datetime import datetime
from io import BytesIO
import os
import openpyxl

from ocr.utils.check_file_exists import check_file_exists
from ocr.utils.extract_document_name import extract_document_name


def create_excel(json_data, pdf_file):
    data_to_insert = {}
    for item in json_data:
        field_temp = {}
        temp = {
            f"code_{item['code']}": field_temp
        }
        for field in item['fields']:
            field_temp.update({field: item['fields'][field]})
        data_to_insert.update(temp)
        
    print(data_to_insert)
            
    excel_dir = f"media/{extract_document_name(str(pdf_file.file))}"
    os.makedirs(excel_dir, exist_ok=True)

    
    file_path = 'ocr/utils/excel/Шаблон для OCR заполненный.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    ws['B3'] = data_to_insert['code_09029']['date']
    ws['C3'] = data_to_insert['code_09029']['number']

    ws['E3'] = data_to_insert['code_04021']['packages lists']
    ws['F3'] = data_to_insert['code_04021']['date']
    ws['G3'] = data_to_insert['code_04021']['terms of deliveries']
    ws['I3'] = data_to_insert['code_04021']['numbers of contacts']
    ws['J3'] = data_to_insert['code_04021']['date contract']

    ws['L3'] = data_to_insert['code_02016']['number']
    ws['M3'] = data_to_insert['code_02016']['date']

    ws['O3'] = data_to_insert['code_01207']['number']
    ws['R3'] = data_to_insert['code_01207']['number']
    ws['P3'] = data_to_insert['code_01207']['date']
    ws['S3'] = data_to_insert['code_01207']['date']

    ws['U3'] = data_to_insert['code_09015']['number']

    ws['W3'] = data_to_insert['code_06999']['date']
    
    excel_file = BytesIO()
    wb.save(excel_file)
    
    excel_file.seek(0)

    return excel_file




def save_excel(file_dir, file_name, wb):
    if check_file_exists(f'{file_dir}/{file_name}.xlsx'):
        dates_path = os.path.join(file_dir, f'{file_name}{datetime.now()}.xlsx')
    else:
        dates_path = os.path.join(file_dir, f'{file_name}.xlsx')

    if dates_path is not None:
        wb.save(dates_path)
        return dates_path
    else:
        print(f"Error: Generated number image is empty for detection")
        return None